# Import Libraries
import requests #Needed For the GET request
import openpyxl #Needed to Open the Excel files
import time #Needed for the sleep time as to not violate the VirusTotal T&C
import xlsxwriter #Needed to write to Excel
import csv #Needed to manipulate the raw data

#Directores for Needed Files
API_KEY_LOCATION = '/Users/conorsherman/Desktop/VT_API_KEY.txt'
TANIUM_RAW_CSV = '/Users/conorsherman/Desktop/Tanium.csv'
TANIUM_TEMP_XLSX = '/Users/conorsherman/Desktop/TEMP_Tanium_Process_Hash.xlsx'
TANIUM_FINAL_XSLX = '/Users/conorsherman/Desktop/Tanium_Output.xlsx'
HASH_CHECKED_FILE = '/Users/conorsherman/Desktop/Tanium_HASH_Checked.xlsx'

#Open Remotely Stored API Key
API_FILE = open(API_KEY_LOCATION, 'r') #This is Unique to You
API_KEY = API_FILE.read().rstrip('\n') #Because, you know... text is "simple"

#Save the RAW .csv as .xlsx
tanium_new_workbook = openpyxl.Workbook()
tanium_new_sheet = tanium_new_workbook.active
tanium_csv = open(TANIUM_RAW_CSV)
reader = csv.reader(tanium_csv, delimiter=',')
for row in reader:
    tanium_new_sheet.append(row)
tanium_new_workbook.save(TANIUM_TEMP_XLSX)

#Open the Excel with List of HASH Previously Checked
hash_checked_workbook = openpyxl.load_workbook(HASH_CHECKED_FILE)
hash_checked_sheet = hash_checked_workbook.worksheets[0]

#Populate the List with items from the HASH Checked Excel
hash_list = []
for row in range(2, hash_checked_sheet.max_row):
    hash_list.append(hash_checked_sheet.cell(row=row, column=2).value)

#Open and assign the Excel
input_workbook = openpyxl.load_workbook(TANIUM_TEMP_XLSX)
input_sheet = input_workbook.worksheets[0]

#Initialize counter user to select value out of the input Excel
#It is Set to 2 so the headers are skipped
input_row = 2

# Temp Dictionary to hold the results
hash_report = {}
# Temo Dictionary to hold the process name
hash_process_name = {}

for hash in range(1, 25): #Needs to Start at 1 due to the headers
    #Populating the HASH and process name
    hash_process_name.update({input_sheet.cell(row=input_row, column=2).value : input_sheet.cell(row=input_row, column=1).value})

    #Get Value of the Cell
    hash_value=input_sheet.cell(row=input_row, column=2).value
    #CHECK There is a HASH Value
    if hash_value:
        #CHECK the HASH is not on the "Checked List"
        if hash_value not in hash_list:
            #Send Hash to VT
            params = {'apikey': API_KEY, 'resource': hash_value}
            headers = {"Accept-Encoding": "gzip, deflate", "User-Agent": "gzip,  ConorSherman"}
            response = requests.get('https://www.virustotal.com/vtapi/v2/file/report', params=params, headers=headers)
            json_response = response.json()

            #Update the Report
            if json_response.get('positives') == None:
                hash_report.update({json_response.get('resource'): "No Results"})
            else:
                hash_report.update({json_response.get('resource'): json_response.get('positives')})

            # Time Delay for Rate Limit
            print "Wait for it...checking input from row %s" %input_row + "\n"
            # Terms and Conditions for VitusTotal
            time.sleep(15)

    else:
        print "ERROR Missing a HASH on row %s" %input_row

    # Move to the next HASH
    input_row += 1




# Prepare Export
# Open the Files to Write To Excel
output_workbook = xlsxwriter.Workbook(TANIUM_FINAL_XSLX)
output_worksheet = output_workbook.add_worksheet()
name_col = 0
hash_col = 1
result_col = 2
output_row = 0

#Write Headers
output_worksheet.write(output_row, name_col, "Process Name")
output_worksheet.write(output_row, hash_col, "Process HASH")
output_worksheet.write(output_row, result_col, "VirusTotal Results")
output_row +=1 #Move to the next row

#Grab the results from the Dictionary
for key, value in hash_report.items():
    output_worksheet.write(output_row, name_col, hash_process_name[key] )
    output_worksheet.write(output_row, hash_col, key) #Write the Hash
    output_worksheet.write(output_row, result_col, hash_report[key]) #Write the Result
    #Position for next Output
    output_row += 1

# Close Excel to keep it clean
output_workbook.close()
